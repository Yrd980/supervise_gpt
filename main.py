import json
import os
from io import BytesIO
from typing import List

import pandas as pd
import uvicorn
from fastapi import FastAPI
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from api import *
from constant import *
from model import RuleObject

app = FastAPI()


@app.post('/upload_file')
async def process(origin_folder: str, target_folder: str):
    # 创建一个任务列表
    tasks = []

    # 遍历源文件夹中的所有文件
    for filename in os.listdir(origin_folder):
        file_path = os.path.join(origin_folder, filename)

        # 检查是否为文件而非目录
        if os.path.isfile(file_path):
            # 以二进制模式打开文件
            with open(file_path, 'rb') as file:
                file_content = file.read()
                # 模拟一个 UploadFile 对象
                file_like = UploadFile(filename=filename, file=BytesIO(file_content))

                rule_result = await upload_file(file_like)

                rule = rule_result.get('ruleList', [])

                # 创建任务并添加到任务列表
                task = asyncio.create_task(process_single_file(filename, rule, target_folder))

                tasks.append(task)
        # 并发执行所有任务
    await asyncio.gather(*tasks)

    return {"message": "All files have been processed."}


async def get_content(rule_list):
    excel_list = []

    for result in rule_list:
        rule_order = result['rule_order']
        rule_content = result['rule_content']
        check_atom_rule_result = await check_atom_rule(rule_content)

        # 如果是复杂条例，拆分为原子条例，否则将原始条例作为单个原子条例
        if check_atom_rule_result['data'] == ClauseType.COMPLEX_CLAUSE.value:
            split_rules_result = await split_atomic_rules(rule_content)
            atom_rules = [rule['atom_rule'] for rule in split_rules_result['ruleList']]
        else:
            atom_rules = [rule_content]

        # 创建 RuleObject 实例并添加到列表中
        rule_obj = RuleObject(
            rule_order=rule_order,
            rule_content=rule_content,
            atom=str(check_atom_rule_result['data']),
            atom_rules=atom_rules
        )
        excel_list.append(rule_obj)

    return excel_list


async def gen_excel(file_name: str, excel_list: List[RuleObject], target_folder: str):
    workbook = Workbook()
    sheet = workbook.active

    headers = ["rule_order", "rule_content", "atom", "atom_rule_content", "Automatable_supervision", "category", "type"]
    sheet.append(headers)

    max_column_width = 40  # 最大列宽度
    max_row_height = 60  # 最大行高度

    async def add_row(sheet_, row_data):
        sheet_.append(row_data)
        await asyncio.sleep(0)  # 让出控制权

    async def process_atom_rules(rule_group_):
        """
        处理原子条例
        """
        rule_order = rule_group_.rule_order
        rule_content = rule_group_.rule_content
        atom = rule_group_.atom
        atom_rules = rule_group_.atom_rules

        if len(atom_rules) > 1:
            first_atom_rule = atom_rules[0]
            super_vise_group = await process_rule(first_atom_rule)
            row_ = [rule_order, rule_content, atom, first_atom_rule, super_vise_group.supervise,
                    super_vise_group.supervise_category, super_vise_group.supervise_type]
            await add_row(sheet, row_)

            start_row = sheet.max_row

            for atom_rule in atom_rules[1:]:
                super_vise_group = await process_rule(atom_rule)
                row_ = ['', '', '', atom_rule, super_vise_group.supervise,
                        super_vise_group.supervise_category, super_vise_group.supervise_type]
                await add_row(sheet, row_)

            end_row = sheet.max_row
            sheet.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            sheet.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
            sheet.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)

            for col in range(1, 4):
                sheet.cell(row=start_row, column=col).alignment = Alignment(horizontal='center', vertical='center')
        else:
            atom_rule = atom_rules[0]
            super_vise_group = await process_rule(atom_rule)
            row_ = [rule_order, rule_content, atom, atom_rule, super_vise_group.supervise,
                    super_vise_group.supervise_category, super_vise_group.supervise_type]
            await add_row(sheet, row_)

    async def worker(queue_):
        while True:
            rule_group_ = await queue_.get()
            if rule_group_ is None:
                break
            await process_atom_rules(rule_group_)
            queue_.task_done()

    queue = asyncio.Queue()
    for rule_group in excel_list:
        await queue.put(rule_group)

    num_workers = 1  # 保证顺序处理
    tasks = []
    for _ in range(num_workers):
        task = asyncio.create_task(worker(queue))
        tasks.append(task)

    await queue.join()

    for _ in range(num_workers):
        await queue.put(None)

    await asyncio.gather(*tasks)

    # 自适应列宽和设置最大宽度
    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        adjusted_length = min(length, max_column_width)  # 取最小值，确保不超过最大宽度
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_length + 5

        # 设置单元格换行和自适应行高，同时设置最大行高
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
                cell_row = cell.row
                current_height = sheet.row_dimensions[cell_row].height or 0
                required_height = 15 * (str(cell.value).count('\n') + 1)
                sheet.row_dimensions[cell_row].height = max(min(max(current_height, required_height), max_row_height),
                                                            15)
        # 确保目标文件夹存在
        if not os.path.exists(target_folder):
            os.makedirs(target_folder)

        # 获取上传文件的原始文件名
        original_filename = file_name
        # 提取文件名，不包括扩展名
        file_name_without_extension = os.path.splitext(original_filename)[0]
        # 创建新的Excel文件名
        excel_filename = f"{file_name_without_extension}.xlsx"

        # 构建完整的文件路径
        file_path = os.path.join(target_folder, excel_filename)

        # 保存工作簿到指定路径，如果文件存在则覆盖
        workbook.save(file_path)


async def process_single_file(file_name: str, rule: List, target_folder: str):
    excel_list = await get_content(rule)
    await gen_excel(file_name, excel_list, target_folder)


@app.post("/count")
async def count(target_folder: str):
    total_count = 0
    # 遍历文件夹并统计
    count_file_path = os.path.join(target_folder, 'count.txt')
    for file_name in os.listdir(target_folder):
        if file_name.endswith('.xlsx'):
            file_path = os.path.join(target_folder, file_name)
            try:
                df = pd.read_excel(file_path)
                tmp = df['Automatable_supervision']
                supervision_count = (tmp == 1).sum()
                total_count += supervision_count
                with open(count_file_path, 'a', encoding='utf-8') as count_file:
                    count_file.write(f"{file_name}: {supervision_count}\n")
            except Exception as e:
                return {"message": f"Error processing file {file_name}: {e}"}
    # 在所有文件处理完毕后，更新 count.txt 文件的总数
    with open(count_file_path, 'r+', encoding='utf-8') as count_file:
        content = count_file.read()
        count_file.seek(0)  # 移动到文件开头
        count_file.write(f"Total Count: {total_count}\n{content}")
    return {"message": "Files processed successfully"}


@app.post('/modify')
async def modify_xlsx(folder: str):
    tasks = []
    for root, _, files in os.walk(folder):
        for file in files:
            if file.lower().endswith('.xlsx'):
                file_path = os.path.join(root, file)
                tasks.append(process_file(file_path))

    results = await asyncio.gather(*tasks)
    return {"message": "Processing complete", "results": results}


def flatten_dict(d, parent_key='', sep='_'):
    items = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep=sep).items())
        elif isinstance(v, list):
            items.append((new_key, json.dumps(v, ensure_ascii=False)))
        else:
            items.append((new_key, str(v)))
    return dict(items)


async def process_file(file_path: str):
    try:
        if not os.path.exists(file_path):
            return {"file": file_path, "status": "Failed", "error": "File does not exist"}
        os.chmod(file_path, 0o666)
        print("File permissions modified successfully!")
        # 读取Excel文件，保留原有格式
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # 读取数据到DataFrame
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]  # 使用第一行作为列名
        df = df.drop(df.index[0])  # 删除重复的列名行
    except PermissionError:
        return {"file": file_path, "status": "Failed to read",
                "error": "Permission denied. The file might be open in another program."}
    except Exception as e:
        return {"file": file_path, "status": "Failed to read", "error": str(e)}

    # 创建新的DataFrame列
    if 'common_element' not in df.columns:
        df['common_element'] = None
    if 'CDSRL_result' not in df.columns:
        df['CDSRL_result'] = None

    # 创建任务列表以并发执行
    process_tasks = []
    for row_index in range(len(df)):
        if df.iloc[row_index]['Automatable_supervision'] == 1:
            process_tasks.append(process_row(df, row_index, df.iloc[row_index]))

    # 并发执行任务
    await asyncio.gather(*process_tasks)

    # 将修改后的数据写回到工作表
    for r_idx, row in enumerate(df.values, start=2):  # 从第2行开始，因为第1行是列名
        for c_idx, value in enumerate(row, start=1):
            if isinstance(value, dict):
                # 如果值是字典，将其扁平化并转换为JSON字符串
                flattened_value = flatten_dict(value)
                cell_value = json.dumps(flattened_value, ensure_ascii=False)
            elif isinstance(value, list):
                # 如果值是列表，直接转换为JSON字符串
                cell_value = json.dumps(value, ensure_ascii=False)
            else:
                cell_value = value
            sheet.cell(row=r_idx, column=c_idx, value=cell_value)
    try:
        workbook.save(file_path)
        return {"file": file_path, "status": "Modified"}
    except Exception as e:
        return {"file": file_path, "status": "Failed to save", "error": str(e)}


async def process_row(df: pd.DataFrame, row_index: int, row: pd.Series):
    try:
        entity_info = await extract_common_element(row['atom_rule_content'], row['category'])
        print(f"Row {row_index}: entity_info", entity_info)

        # 使用 iloc 按位置访问行
        df.iloc[row_index, df.columns.get_loc('common_element')] = json.dumps(entity_info, ensure_ascii=False)

        cdsrl_result = await generate_cdsrl(row['atom_rule_content'], row['category'], entity_info)
        print(f"Row {row_index}: cdsrl_result", cdsrl_result)
        print("========================")

        # 使用 iloc 按位置访问行
        df.iloc[row_index, df.columns.get_loc('CDSRL_result')] = json.dumps(cdsrl_result, ensure_ascii=False)
    except Exception as e:
        print(f"Error processing row {row_index}: {str(e)}")
        # 可以选择在这里设置一个错误值
        df.iloc[row_index, df.columns.get_loc('common_element')] = json.dumps({"error": str(e)}, ensure_ascii=False)
        df.iloc[row_index, df.columns.get_loc('CDSRL_result')] = json.dumps({"error": str(e)}, ensure_ascii=False)


def process_excel(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # 检查是否已存在这两列，如果不存在则添加
    if 'common_element' not in sheet[1]:
        sheet.cell(row=1, column=sheet.max_column - 1, value='common_element')
    if 'CDSRL_result' not in sheet[1]:
        sheet.cell(row=1, column=sheet.max_column, value='CDSRL_result')

    # 获取这两列的列号
    common_element_col = None
    cdsrl_result_col = None
    for cell in sheet[1]:
        if cell.value == 'common_element':
            common_element_col = cell.column
        elif cell.value == 'CDSRL_result':
            cdsrl_result_col = cell.column

    if common_element_col and cdsrl_result_col:
        for col in [common_element_col, cdsrl_result_col]:
            letter = get_column_letter(col)
            max_length = 0
            for cell in sheet[letter]:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[letter].width = adjusted_width

            for cell in sheet[letter]:
                cell.alignment = Alignment(vertical='bottom', wrap_text=True)

    workbook.save(file_path)


@app.post("/beautify")
async def beautify_excel_files(folder_path: str):
    if not os.path.isdir(folder_path):
        from fastapi import HTTPException
        raise HTTPException(status_code=400, detail="Invalid folder path")

    processed_files = []
    for filename in os.listdir(folder_path):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(folder_path, filename)
            try:
                process_excel(file_path)
                processed_files.append(filename)
            except Exception as e:
                print(f"Error processing {filename}: {str(e)}")

    return {"message": f"Processed {len(processed_files)} files", "processed_files": processed_files}


def set_column_width(file_path: str, width: float = 15):
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # 获取最后两列的列号
    last_column = sheet.max_column
    second_last_column = last_column - 1

    # 将厘米转换为Excel单位（约4.7个字符宽度等于1厘米）
    width_excel_units = width

    # 设置最后两列的宽度
    for col in [second_last_column, last_column]:
        column_letter = get_column_letter(col)
        sheet.column_dimensions[column_letter].width = width_excel_units

    workbook.save(file_path)


@app.post("/set_column_width")
async def set_excel_column_width(folder_path: str, width: float = 30):
    if not os.path.isdir(folder_path):
        from fastapi import HTTPException
        raise HTTPException(status_code=400, detail="Invalid folder path")

    processed_files = []
    for filename in os.listdir(folder_path):
        if filename.endswith(('.xlsx', '.xls')):
            file_path = os.path.join(folder_path, filename)
            try:
                set_column_width(file_path, width)
                processed_files.append(filename)
            except Exception as e:
                print(f"Error processing {filename}: {str(e)}")

    return {"message": f"Processed {len(processed_files)} files", "processed_files": processed_files}


if __name__ == '__main__':
    uvicorn.run(app, host='0.0.0.0', port=8000)
